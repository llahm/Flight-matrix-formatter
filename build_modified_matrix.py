"""
build_modified_matrix.py

Reshapes a "TRANSFER MATRIX" sheet (incoming flights as columns, outbound
flights as rows, cell = pax/pieces transferring) into the "modified matrix"
layout: one 2-column block per incoming flight (Destination, Count),
non-zero only, with each flight's own total at the bottom of its block.
Destinations are listed in departure-time order by default (same order as
the source sheet), or by transfer count with --sort-by count.

Usage (from command line):
    python3 build_modified_matrix.py SOURCE.xlsx OUTPUT.xlsx \
        [--sheet "SHEET NAME"] [--sort-by departure|count] \
        [--min-sta HH:MM] [--max-sta HH:MM] [--outbound CODE1,CODE2,...]

If --sheet is left off, the workbook's active/only sheet is used - so a
single-tab file needs no sheet name at all. Pass --sheet ALL to process
every matrix block on every tab in one go (OUTPUT is then a directory).

Or import generate_modified_matrix() and call it directly / in a loop over
several sheets (dates).
"""
import sys
import argparse
import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"


def _parse_time(s):
    if s is None:
        return None
    h, m = s.split(":")
    return datetime.time(int(h), int(m))


def find_matrix_blocks(ws):
    """Some tabs hold two stacked matrices (morning + evening) rather than
    one. Scan column A for every cell containing 'TRANSFER MATRIX' and
    return the row each one starts on, so each block can be processed
    separately."""
    rows = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "TRANSFER MATRIX" in v.upper():
            rows.append(r)
    return rows


def _read_source(ws, title_row=1):
    """Pull the incoming-flight headers and the outbound rows out of one
    transfer-matrix block, starting at title_row. Fully driven off the
    sheet's own extent (relative to title_row), so it works on any block,
    on any day's tab, that follows the same layout:
        title_row + 0 : title
        title_row + 1 : 'City' / incoming flight codes (row2 in single-block sheets)
        title_row + 2 : 'FLT#' / 'InBound' / incoming flight numbers
        title_row + 3 : 'OutBound' / 'STA' / incoming STA times
        title_row + 4 : 'STD/ETA' / incoming STD/ETA times
        title_row + 5 : blank
        title_row + 6.. : outbound rows, until a 'Total' cell in col A
    """
    title = ws.cell(row=title_row, column=1).value
    code_row = title_row + 1
    flt_row = title_row + 2
    sta_row = title_row + 3
    eta_row = title_row + 4
    data_row0 = title_row + 6

    # Incoming flights run across columns D.. (col 4)
    incoming = []
    col = 4
    while ws.cell(row=code_row, column=col).value is not None:
        incoming.append({
            "col": col,
            "code": ws.cell(row=code_row, column=col).value,
            "flt": ws.cell(row=flt_row, column=col).value,
            "sta": ws.cell(row=sta_row, column=col).value,
            "std_eta": ws.cell(row=eta_row, column=col).value,
        })
        col += 1

    # Outbound rows run down column A from data_row0 until the 'Total' row
    # or the next matrix's title row, whichever comes first.
    outbound = []
    row = data_row0
    while True:
        code = ws.cell(row=row, column=1).value
        if code is None or str(code).strip().lower() == "total":
            break
        if isinstance(code, str) and "TRANSFER MATRIX" in code.upper():
            break
        outbound.append({
            "row": row,
            "code": code,
            "flt": ws.cell(row=row, column=2).value,
            "sta": ws.cell(row=row, column=3).value,
        })
        row += 1

    return title, incoming, outbound


def generate_modified_matrix(
    source_path,
    sheet_name=None,
    output_path="modified_matrix.xlsx",
    min_sta=None,
    max_sta=None,
    outbound_codes=None,
    sort_by="departure",
    title_row=1,
    src_wb=None,
):
    """
    source_path, sheet_name : the workbook + tab to read (e.g. the daily
        TRANSFER MATRIX sheet). Leave sheet_name as None to use the
        workbook's active/only sheet - handy for a file that has just one
        tab and you don't want to type its name.
    output_path : where to write the reshaped workbook.
    min_sta, max_sta : datetime.time (or 'HH:MM' string) bounds on the
        INCOMING flight's arrival (STA). Only incoming flights whose STA
        falls in [min_sta, max_sta] are included. Either can be left None.
    outbound_codes : optional list of destination codes (e.g.
        ['IST', 'IAD', 'ORD']) to restrict which OUTBOUND rows are
        considered at all. None = include every outbound row.
    sort_by : how each flight's destination list is ordered.
        'departure' (default) - keeps the same order the destinations
            appear in the source sheet, which is sorted by outbound
            departure time ascending, exactly like the original workbook.
        'count' - sorts by transfer count, largest first.
    title_row : row the 'TRANSFER MATRIX ...' title sits on for this block.
        Most tabs have exactly one matrix starting at row 1. Some tabs hold
        two stacked matrices (morning + evening) - use
        find_matrix_blocks(ws) to get their title rows, or use
        generate_all() below to handle this automatically.
    src_wb : optionally pass an already-open workbook (data_only=True) to
        avoid re-reading the source file for every block/sheet.
    """
    if isinstance(min_sta, str):
        min_sta = _parse_time(min_sta)
    if isinstance(max_sta, str):
        max_sta = _parse_time(max_sta)
    outbound_set = set(outbound_codes) if outbound_codes else None
    if sort_by not in ("departure", "count"):
        raise ValueError("sort_by must be 'departure' or 'count'")

    if src_wb is None:
        src_wb = openpyxl.load_workbook(source_path, data_only=True)
    if sheet_name is None:
        ws = src_wb.active  # the workbook's only/currently-active tab
    else:
        ws = src_wb[sheet_name]
    title, incoming, outbound = _read_source(ws, title_row=title_row)

    if outbound_set is not None:
        outbound = [o for o in outbound if o["code"] in outbound_set]

    def sta_in_range(sta):
        if sta is None:
            return False
        if min_sta is not None and sta < min_sta:
            return False
        if max_sta is not None and sta > max_sta:
            return False
        return True

    if min_sta is not None or max_sta is not None:
        incoming = [f for f in incoming if sta_in_range(f["sta"])]

    # Pull the counts for each surviving incoming flight, keep non-zero only.
    # A handful of source cells use a "n*m" annotation (e.g. "1*1") instead
    # of a plain number; keep the original label but sort on its numeric
    # value (the product).
    def numeric_value(val):
        if isinstance(val, (int, float)):
            return val
        if isinstance(val, str) and "*" in val:
            try:
                parts = [float(p) for p in val.split("*")]
                out = 1.0
                for p in parts:
                    out *= p
                return out
            except ValueError:
                return 0
        try:
            return float(val)
        except (TypeError, ValueError):
            return 0

    for f in incoming:
        pairs = []
        for o in outbound:
            val = ws.cell(row=o["row"], column=f["col"]).value
            if val not in (None, 0, "0"):
                pairs.append((o["code"], val, numeric_value(val)))
        # 'departure' order = the order outbound rows already appear in the
        # source (chronological by departure time) - just leave it as-is.
        # 'count' order = re-sort by transfer count, largest first.
        if sort_by == "count":
            pairs.sort(key=lambda p: p[2], reverse=True)
        f["pairs"] = [(dest, val) for dest, val, _ in pairs]

    # ---- write output ----
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Modified Matrix"

    bold = Font(name=FONT_NAME, bold=True)
    normal = Font(name=FONT_NAME)

    if title:
        out_ws.cell(row=1, column=1, value=f"MODIFIED MATRIX - {title}").font = bold

    HEADER_ROW_CODE = 3
    HEADER_ROW_FLT = 4
    HEADER_ROW_STA = 5
    HEADER_ROW_ETA = 6
    DATA_START_ROW = 8

    block_col = 2  # start at column B, like the example (leaves col A clear)
    for f in incoming:
        dest_col = block_col
        cnt_col = block_col + 1

        out_ws.cell(row=HEADER_ROW_CODE, column=dest_col, value=f["code"]).font = bold
        out_ws.cell(row=HEADER_ROW_FLT, column=dest_col, value=f["flt"]).font = normal
        c = out_ws.cell(row=HEADER_ROW_STA, column=dest_col, value=f["sta"])
        c.font = normal
        c.number_format = "hh:mm"
        c = out_ws.cell(row=HEADER_ROW_ETA, column=dest_col, value=f["std_eta"])
        c.font = normal
        c.number_format = "hh:mm"

        r = DATA_START_ROW
        for dest, cnt in f["pairs"]:
            out_ws.cell(row=r, column=dest_col, value=dest).font = normal
            cell = out_ws.cell(row=r, column=cnt_col, value=numeric_value(cnt))
            cell.font = normal
            if isinstance(cnt, str) and "*" in cnt:
                cell.comment = openpyxl.comments.Comment(
                    f"Source cell read '{cnt}'; value shown is the product.",
                    "build_modified_matrix.py",
                )
            r += 1

        # Real formula total for THIS flight's own block only
        if f["pairs"]:
            total_row = r
            col_letter = get_column_letter(cnt_col)
            tf = out_ws.cell(
                row=total_row, column=cnt_col,
                value=f"=SUM({col_letter}{DATA_START_ROW}:{col_letter}{total_row-1})"
            )
            tf.font = bold
            out_ws.cell(row=total_row, column=dest_col, value="Total").font = bold

        block_col += 3  # 2 data cols + 1 spacer

    # column widths
    for col in range(2, block_col):
        out_ws.column_dimensions[get_column_letter(col)].width = 10

    out_wb.save(output_path)
    return output_path


def generate_all(
    source_path,
    output_dir,
    sheet_names=None,
    min_sta=None,
    max_sta=None,
    outbound_codes=None,
    sort_by="departure",
):
    """
    Run every matrix block on every requested sheet in one go.

    sheet_names : list of tab names to process, or None for every sheet in
        the workbook that contains at least one 'TRANSFER MATRIX' block.
    Everything else (min_sta/max_sta/outbound_codes) is applied identically
    to every block. Returns the list of output file paths written.
    """
    import os

    src_wb = openpyxl.load_workbook(source_path, data_only=True)
    if sheet_names is None:
        sheet_names = src_wb.sheetnames

    written = []
    for sn in sheet_names:
        ws = src_wb[sn]
        title_rows = find_matrix_blocks(ws)
        for title_row in title_rows:
            title, _, _ = _read_source(ws, title_row=title_row)
            # Build a filesystem-safe name from the block's own title,
            # e.g. "TRANSFER MATRIX FOR THE DATE 28-JUN-26/EVENING"
            safe = "".join(ch if ch.isalnum() else "_" for ch in (title or f"{sn}_{title_row}"))
            safe = "_".join(filter(None, safe.split("_")))
            out_path = os.path.join(output_dir, f"MODIFIED_{safe}.xlsx")
            generate_modified_matrix(
                source_path, sn, out_path,
                min_sta=min_sta, max_sta=max_sta, outbound_codes=outbound_codes,
                sort_by=sort_by, title_row=title_row, src_wb=src_wb,
            )
            written.append(out_path)
    return written


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("source", help="Source .xlsx (the transfer matrix workbook)")
    ap.add_argument("output", help="Output .xlsx path, or an output DIRECTORY when --sheet ALL is used")
    ap.add_argument("--sheet", default=None,
                     help="Tab name to read, e.g. \"SUN INTNL\". Pass 'ALL' to process every matrix "
                          "block on every tab. Leave this off entirely if the workbook only has one "
                          "sheet (or you just want whichever sheet is currently active) - it will be "
                          "picked automatically.")
    ap.add_argument("--sort-by", choices=["departure", "count"], default="departure",
                     help="'departure' (default) keeps each flight's destinations in the same "
                          "departure-time order as the source sheet. 'count' sorts by transfer "
                          "count, largest first.")
    ap.add_argument("--title-row", type=int, default=1,
                     help="Row the 'TRANSFER MATRIX...' title is on, for tabs with more than one stacked matrix. Ignored when --sheet ALL.")
    ap.add_argument("--min-sta", default=None, help="HH:MM, only incoming flights with STA at/after this time")
    ap.add_argument("--max-sta", default=None, help="HH:MM, only incoming flights with STA at/before this time")
    ap.add_argument("--outbound", default=None, help="Comma-separated outbound codes to restrict to, e.g. IST,IAD,ORD")
    args = ap.parse_args()

    outbound_codes = args.outbound.split(",") if args.outbound else None

    if args.sheet is not None and args.sheet.upper() == "ALL":
        import os
        os.makedirs(args.output, exist_ok=True)
        paths = generate_all(
            args.source, args.output,
            min_sta=args.min_sta, max_sta=args.max_sta,
            outbound_codes=outbound_codes, sort_by=args.sort_by,
        )
        print(f"Wrote {len(paths)} file(s):")
        for p in paths:
            print(" ", p)
    else:
        generate_modified_matrix(
            args.source, args.sheet, args.output,
            min_sta=args.min_sta, max_sta=args.max_sta,
            outbound_codes=outbound_codes, sort_by=args.sort_by,
            title_row=args.title_row,
        )
        print(f"Wrote {args.output}")
